"""
Excel file handler for reading and writing transaction data.
"""

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class XlsxHandler:
    """Handler for Excel file operations."""

    # Column headers in order
    HEADERS = [
        "Transaction Hash",
        "Blockno",
        "UnixTimestamp",
        "DateTime (UTC)",
        "From",
        "To",
        "TokenValue",
        "USDValueDayOfTx",
        "ContractAddress",
        "TokenName",
        "TokenSymbol"
    ]

    # Column widths for better readability
    COLUMN_WIDTHS = {
        "Transaction Hash": 70,
        "Blockno": 12,
        "UnixTimestamp": 14,
        "DateTime (UTC)": 14,
        "From": 45,
        "To": 45,
        "TokenValue": 20,
        "USDValueDayOfTx": 16,
        "ContractAddress": 45,
        "TokenName": 25,
        "TokenSymbol": 12
    }

    def __init__(self, file_path: str):
        """
        Initialize the Excel handler.

        Args:
            file_path: Path to the Excel file
        """
        self.file_path = Path(file_path)

    def file_exists(self) -> bool:
        """Check if the file already exists."""
        return self.file_path.exists()

    def create_new_file(self) -> None:
        """Create a new Excel file with headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Write headers
        for col, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            # Make headers bold
            cell.font = cell.font.copy(bold=True)

        # Set column widths
        for col, header in enumerate(self.HEADERS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = self.COLUMN_WIDTHS.get(header, 15)

        # Freeze the header row
        ws.freeze_panes = "A2"

        wb.save(self.file_path)

    def get_last_row(self) -> int:
        """
        Get the last row number with data.

        Returns:
            Last row number (1 if only headers, 0 if file doesn't exist)
        """
        if not self.file_exists():
            return 0

        wb = load_workbook(self.file_path, read_only=True)
        ws = wb.active
        last_row = ws.max_row
        wb.close()
        return last_row

    def get_existing_hashes(self) -> set[str]:
        """
        Get all existing transaction hashes to avoid duplicates.

        Returns:
            Set of transaction hashes already in the file
        """
        if not self.file_exists():
            return set()

        hashes = set()
        wb = load_workbook(self.file_path, read_only=True)
        ws = wb.active

        # Find the Transaction Hash column (should be column 1)
        hash_col = 1
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                hashes.add(row[0])

        wb.close()
        return hashes

    def get_last_timestamp(self) -> int | None:
        """
        Get the last (most recent) Unix timestamp from existing data.

        Returns:
            Last Unix timestamp or None if no data
        """
        if not self.file_exists():
            return None

        wb = load_workbook(self.file_path, read_only=True)
        ws = wb.active

        last_timestamp = None
        # UnixTimestamp is column 3
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
            if row[0]:
                try:
                    timestamp = int(row[0])
                    if last_timestamp is None or timestamp > last_timestamp:
                        last_timestamp = timestamp
                except (ValueError, TypeError):
                    continue

        wb.close()
        return last_timestamp

    def append_transactions(self, transactions: list[dict]) -> int:
        """
        Append transactions to the Excel file.
        Will not modify existing data, formulas, or formatting.
        Skips duplicate transactions based on hash.

        Args:
            transactions: List of transaction dictionaries

        Returns:
            Number of transactions actually added (excluding duplicates)
        """
        if not transactions:
            return 0

        # Create file if it doesn't exist
        if not self.file_exists():
            self.create_new_file()

        # Get existing hashes to avoid duplicates
        existing_hashes = self.get_existing_hashes()

        # Filter out duplicates
        new_transactions = [
            tx for tx in transactions
            if tx.get("Transaction Hash") not in existing_hashes
        ]

        if not new_transactions:
            return 0

        # Load workbook (not read-only so we can write)
        wb = load_workbook(self.file_path)
        ws = wb.active

        # Find the next empty row
        next_row = ws.max_row + 1

        # Append each transaction - values only, no formatting applied
        for tx in new_transactions:
            for col, header in enumerate(self.HEADERS, start=1):
                value = tx.get(header, "")
                # Just write the value, don't touch formatting at all
                ws.cell(row=next_row, column=col, value=value)
            next_row += 1

        wb.save(self.file_path)
        return len(new_transactions)

    def get_row_count(self) -> int:
        """
        Get the number of data rows (excluding header).

        Returns:
            Number of data rows
        """
        last_row = self.get_last_row()
        return max(0, last_row - 1)  # Subtract header row
